"""Export entry points for o-grid datasets."""

from __future__ import annotations

from collections.abc import Iterable, Sequence
from pathlib import Path
from typing import Any

from infrasys import System
from loguru import logger
from openpyxl import Workbook
from openpyxl.cell.cell import Cell

from o_grid.acpf.models.results import PowerFlowResults
from o_grid.constants import DEFAULT_SEPARATOR, SHEET_HEADERS
from o_grid.utils.utils_exporter import format_rows

SheetRows = tuple[Sequence[str], Iterable[Sequence[object]]]


class ExportSolution:
    """Export solved power-flow components using the reference result schema."""

    def __init__(
        self,
        *,
        system: System,
        format: str,
        output_path: str | Path,
        export: bool = True,
    ) -> None:
        export_format = format.strip().lower()
        if export_format not in {"excel", "xlsx"}:
            raise ValueError(f"Unsupported solution export format: {format!r}")
        results = getattr(system, "power_flow_results", None)
        if not isinstance(results, PowerFlowResults):
            raise ValueError("The system does not contain solved power-flow results")

        self.system = system
        self.format = "excel"
        self.output_path = Path(output_path)
        self.export = export
        if not export:
            return
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        _write_excel(results, self.output_path)
        logger.info("ExportSolution saved on {}", self.output_path)


def export_rows(rows: list[dict[str, object]], separator: str = DEFAULT_SEPARATOR) -> str:
    """Serialize rows into delimited text."""
    return format_rows(rows, separator=separator)


def _write_excel(results: PowerFlowResults, output_path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, (headers, rows) in _workbook_rows(results, output_path).items():
        sheet = workbook.create_sheet(sheet_name)
        sheet.freeze_panes = "A2"
        sheet.append(headers)
        for row in rows:
            sheet.append(tuple(_excel_value(value) for value in row))
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                _format_number(cell)
    workbook.save(output_path)


def _workbook_rows(results: PowerFlowResults, output_path: Path) -> dict[str, SheetRows]:
    buses = results.ac_buses
    information = results.information
    generators = results.generators
    loads = [
        bus for bus in buses if abs(bus.active_load_mw) > 1e-9 or abs(bus.reactive_load_mvar) > 1e-9
    ]
    summary = (
        ("Case", information.source_path),
        ("Workbook", str(output_path)),
        ("Method", information.method or information.solver),
        ("Converged", information.converged),
        ("Iterations", information.iterations),
        ("Max Mismatch (pu)", information.max_mismatch_pu),
        ("Base MVA", information.base_mva),
        ("Buses", information.bus_count),
        ("Buses After Reduction", information.bus_count_after_reduction),
        ("Generators", len(generators)),
        ("Loads", len(loads)),
        ("Lines", information.branch_count),
        ("Lines After Reduction", information.branch_count_after_reduction),
        ("Transmission Lines", len(results.ac_lines)),
        ("Transformers", len(results.transformers)),
        ("LTC", len(results.ltc_transformers)),
        ("PST", len(results.phase_shifting_transformers)),
        ("HVDC", information.dc_line_count),
        ("SVC", len(results.static_var_compensators)),
        ("CSC", len(results.controllable_series_compensators)),
        ("Voltage Upper Violations", information.voltage_upper_violations),
        ("Voltage Lower Violations", information.voltage_lower_violations),
        ("Line Flow Overloads", information.line_flow_overloads),
    )
    rows: dict[str, Iterable[Sequence[object]]] = {
        "Summary": summary,
        "Buses": (
            (
                bus.bus_number,
                bus.bus_name,
                bus.bus_type,
                bus.area,
                bus.in_service,
                bus.voltage_pu,
                bus.voltage_kv,
                bus.angle_deg,
                bus.active_generation_mw,
                bus.reactive_generation_mvar,
                bus.active_load_mw,
                bus.reactive_load_mvar,
                bus.minimum_voltage_pu,
                bus.maximum_voltage_pu,
                bus.violation or "",
                bus.representative_bus,
                bus.collapsed,
            )
            for bus in buses
        ),
        "Generators": (
            (
                item.bus_number,
                item.bus_name,
                item.generator_type,
                item.active_generation_mw,
                item.reactive_generation_mvar,
                item.maximum_active_generation_mw,
                item.reserve_mw,
                item.voltage_pu,
                item.angle_deg,
            )
            for item in generators
        ),
        "Loads": (
            (bus.bus_number, bus.bus_name, bus.active_load_mw, bus.reactive_load_mvar, "Bus")
            for bus in loads
        ),
        "Lines": (
            (
                line.line_number,
                line.from_bus,
                line.to_bus,
                line.circuit,
                line.resistance_pu,
                line.reactance_pu,
                line.charging_pu,
                line.tap_pu,
                line.phase_shift_deg,
                line.rating_mva,
                line.active_from_mw,
                line.reactive_from_mvar,
                line.power_factor_from,
                line.reactive_type_from,
                line.active_to_mw,
                line.reactive_to_mvar,
                line.power_factor_to,
                line.reactive_type_to,
                line.loading_percent,
                line.active_loss_mw,
                line.reactive_loss_mvar,
                "true" if line.violation else "false",
            )
            for line in results.ac_lines
        ),
        "Transformers": (
            (
                item.device_number,
                item.from_bus,
                item.to_bus,
                item.circuit,
                item.resistance_pu,
                item.reactance_pu,
                item.tap_pu,
                item.phase_shift_deg,
                item.rating_mva,
                item.active_from_mw,
                item.reactive_from_mvar,
                item.power_factor_from,
                item.reactive_type_from,
                item.active_to_mw,
                item.reactive_to_mvar,
                item.power_factor_to,
                item.reactive_type_to,
                item.loading_percent,
                item.active_loss_mw,
                item.reactive_loss_mvar,
                "true" if item.violation else "false",
            )
            for item in results.transformers
        ),
        "LTC": (
            (
                item.device_number,
                item.from_bus,
                item.to_bus,
                item.circuit,
                item.controlled_bus,
                item.tap_pu,
                item.minimum_tap_pu,
                item.maximum_tap_pu,
                item.target_voltage_pu,
                item.active_from_mw,
                item.reactive_from_mvar,
                item.power_factor_from,
                item.reactive_type_from,
                item.active_to_mw,
                item.reactive_to_mvar,
                item.power_factor_to,
                item.reactive_type_to,
            )
            for item in results.ltc_transformers
        ),
        "PST": (
            (
                item.device_number,
                item.from_bus,
                item.to_bus,
                item.circuit,
                item.controlled_bus,
                item.phase_shift_deg,
                item.minimum_phase_shift_deg,
                item.maximum_phase_shift_deg,
                item.target_active_power_mw,
                item.active_from_mw,
                item.reactive_from_mvar,
                item.power_factor_from,
                item.reactive_type_from,
                item.active_to_mw,
                item.reactive_to_mvar,
                item.power_factor_to,
                item.reactive_type_to,
            )
            for item in results.phase_shifting_transformers
        ),
        "HVDC": (
            (
                item.bus_number,
                item.bus_name,
                item.voltage_pu,
                item.converter_type,
                item.pole_number,
                item.control_mode,
                item.active_power_mw,
                item.reactive_power_mvar,
                item.loss_mw,
                item.dc_voltage_kv,
                item.dc_current_pu,
                item.dc_current_a,
                item.firing_angle_deg,
                item.overlap_angle_deg,
                item.power_factor_angle_deg or "-",
                item.tap_pu,
                item.status,
            )
            for item in results.dc_lines
        ),
        "SVC": (
            (
                item.device_number,
                item.bus_number,
                item.controlled_bus,
                item.mode,
                item.reactive_power_mvar,
                item.minimum_reactive_power_mvar,
                item.maximum_reactive_power_mvar,
                item.equation_residual,
                _svc_state(item),
            )
            for item in results.static_var_compensators
        ),
        "CSC": (
            (
                item.device_number,
                item.from_bus,
                item.to_bus,
                item.circuit,
                item.mode,
                item.reactance_pu,
                item.minimum_reactance_pu,
                item.maximum_reactance_pu,
                item.active_from_mw,
                item.reactive_from_mvar,
                item.power_factor_from,
                item.reactive_type_from,
                item.active_to_mw,
                item.reactive_to_mvar,
                item.power_factor_to,
                item.reactive_type_to,
                item.status,
            )
            for item in results.controllable_series_compensators
        ),
    }
    return {name: (SHEET_HEADERS[name], rows[name]) for name in SHEET_HEADERS}


def _svc_state(item: Any) -> str:
    if item.status != "InSvc":
        return item.status
    if (
        item.minimum_reactive_power_mvar is not None
        and item.reactive_power_mvar <= item.minimum_reactive_power_mvar + 1e-6
    ):
        return "Qmin"
    if (
        item.maximum_reactive_power_mvar is not None
        and item.reactive_power_mvar >= item.maximum_reactive_power_mvar - 1e-6
    ):
        return "Qmax"
    return "Free"


def _excel_value(value: object) -> object:
    return "" if value is None else value


def _format_number(cell: Cell) -> None:
    if isinstance(cell.value, float):
        cell.number_format = "0.0000"
