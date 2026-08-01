from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from o_grid import ExportSolution
from o_grid.acpf import NewtonRaphsonPowerFlow, OptimizationACPowerFlow
from o_grid.exporter import SHEET_HEADERS, _svc_state
from o_grid.parser import AnaredeInfrasysParser

DATA = Path(__file__).parent / "data" / "pwf"


class _Svc:
    def __init__(
        self, status: str, q: float, qmin: float | None = None, qmax: float | None = None
    ) -> None:
        self.status = status
        self.reactive_power_mvar = q
        self.minimum_reactive_power_mvar = qmin
        self.maximum_reactive_power_mvar = qmax


@pytest.mark.parametrize(
    ("svc", "expected"),
    [
        (_Svc("OutSvc", 0.0), "OutSvc"),
        (_Svc("InSvc", 10.0, qmin=10.0, qmax=50.0), "Qmin"),
        (_Svc("InSvc", 50.0, qmin=10.0, qmax=50.0), "Qmax"),
        (_Svc("InSvc", 30.0, qmin=10.0, qmax=50.0), "Free"),
        (_Svc("InSvc", 0.0), "Free"),
    ],
)
def test_svc_state(svc: _Svc, expected: str) -> None:
    assert _svc_state(svc) == expected


def test_export_solution_writes_reference_excel_schema(tmp_path: Path) -> None:
    parsed = AnaredeInfrasysParser().parse(DATA / "d_9nodes.pwf")
    solved = NewtonRaphsonPowerFlow(parsed.system)
    output_path = tmp_path / "solution.xlsx"

    export = ExportSolution(system=solved, format="excel", output_path=output_path)

    workbook = load_workbook(output_path, data_only=True)
    assert export.system is solved
    assert export.output_path == output_path
    assert workbook.sheetnames == list(SHEET_HEADERS)
    for sheet_name, expected_headers in SHEET_HEADERS.items():
        sheet = workbook[sheet_name]
        assert tuple(cell.value for cell in sheet[1]) == expected_headers
        assert sheet.freeze_panes == "A2"
    summary = dict(workbook["Summary"].values)
    assert summary["Method"] == "NewtonRaphsonPowerFlow"
    assert summary["Converged"] is True
    assert summary["Buses"] == 9
    assert summary["Buses After Reduction"] <= summary["Buses"]
    assert summary["Lines After Reduction"] <= summary["Lines"]
    assert workbook["Buses"].max_row == 10
    assert workbook["Generators"].max_row == 3
    assert "Transformers" in workbook.sheetnames
    generator_headers = tuple(cell.value for cell in workbook["Generators"][1])
    assert "Type" in generator_headers
    assert "Reserve(MW)" in generator_headers
    for sheet_name in ("Lines", "Transformers", "LTC", "PST", "CSC"):
        headers = tuple(cell.value for cell in workbook[sheet_name][1])
        assert "Power Factor:From-To" in headers
        assert "F:Capac/Inductive" in headers
        assert "Power Factor:To-From" in headers
        assert "T:Capac/Inductive" in headers
    assert workbook["Buses"]["F2"].number_format == "0.0000"


def test_export_solution_requires_solved_system(tmp_path: Path) -> None:
    parsed = AnaredeInfrasysParser().parse(DATA / "d_9nodes.pwf")

    with pytest.raises(ValueError, match="does not contain solved"):
        ExportSolution(
            system=parsed.system,
            format="excel",
            output_path=tmp_path / "solution.xlsx",
        )


def test_export_solution_rejects_unknown_format(tmp_path: Path) -> None:
    parsed = AnaredeInfrasysParser().parse(DATA / "d_9nodes.pwf")
    solved = NewtonRaphsonPowerFlow(parsed.system)

    with pytest.raises(ValueError, match="Unsupported solution export format"):
        ExportSolution(system=solved, format="csv", output_path=tmp_path / "solution.csv")


def test_export_summary_reports_optimization_method(tmp_path: Path) -> None:
    parsed = AnaredeInfrasysParser().parse(DATA / "d_9nodes.pwf")
    solved = OptimizationACPowerFlow(parsed.system)
    output_path = tmp_path / "solution.xlsx"

    ExportSolution(system=solved, format="excel", output_path=output_path)

    workbook = load_workbook(output_path, data_only=True)
    summary = dict(workbook["Summary"].values)
    assert summary["Method"] == "OptimizationACPowerFlow"
    assert summary["Converged"] is True
