from __future__ import annotations

from pathlib import Path

import pytest
from infrasys import System
from openpyxl import load_workbook
from r2x_core.plugin_context import PluginContext

from o_grid.acpf import (  # noqa: F401
    FastDecoupledPowerFlow,
    NewtonRaphsonPowerFlow,
    OptimizationACPowerFlow,
)
from o_grid.acpf.models.lcc import build_lcc_data
from o_grid.exporter import ExportSolution
from o_grid.matpower import MatpowerInfrasysParser, parse_matpower_system
from o_grid.models import (
    ACBus,
    ACBusTypes,
    ACLine,
    Area,
    Generator,
    PhaseShiftingTransformer,
    SwitchDevice,
    TransformerDevice,
)
from o_grid.models.m_models import Branch, branch_block
from o_grid.plugin_config import AnaredeConfig, MatpowerConfig
from o_grid.plugin_parser import MatPowerParser
from o_grid.units import get_magnitude

MATPOWER_DATA = Path(__file__).parent / "data" / "mat"


def test_parse_case5_dc_components() -> None:
    parsed = parse_matpower_system(MATPOWER_DATA / "case5_dc.m", system_name="case5_dc")

    assert parsed.system.name == "case5_dc"
    assert len(parsed.components_by_block["DBAR"]) == 5
    assert len(parsed.components_by_block["DLIN"]) == 6
    assert len(parsed.components_by_block["DGER"]) == 5
    assert len(parsed.components_by_block["DCTE"]) == 1

    buses = {bus.number: bus for bus in parsed.components_by_block["DBAR"]}
    assert buses[1].bustype == ACBusTypes.PQ
    assert buses[4].bustype == ACBusTypes.REF
    assert get_magnitude(buses[3].active_generation) == pytest.approx(333.6866, rel=1e-4)
    assert get_magnitude(buses[1].min_reactive_generation) == pytest.approx(-157.5)
    assert buses[2].min_reactive_generation is None

    assert len(list(parsed.system.get_components(Area))) == 1
    assert len(list(parsed.system.get_components(Generator))) == 5
    assert len(list(parsed.system.get_components(ACBus))) == 5


def test_parse_case5_dc_dc_records() -> None:
    parsed = parse_matpower_system(MATPOWER_DATA / "case5_dc.m")
    components = parsed.components_by_block
    assert len(components["DCBA"]) == 2
    assert len(components["DCNV"]) == 2
    assert len(components["DCCV"]) == 2
    assert len(components["DELO"]) == 1
    assert len(components["DCLI"]) == 1

    lccs = build_lcc_data(components)
    assert len(lccs) == 1
    lcc = lccs[0]
    assert lcc.rectifier_bus == 3
    assert lcc.inverter_bus == 5
    assert lcc.rectifier_slack is False
    assert lcc.inverter_slack is True
    assert lcc.pdc_mw == pytest.approx(15.0)
    assert lcc.rdc_ohm == pytest.approx(1.0)
    assert lcc.alpha_deg == pytest.approx(15.0)
    assert lcc.gamma_deg == pytest.approx(18.0)


def test_parse_case7_tplgy_status_handling() -> None:
    parsed = parse_matpower_system(MATPOWER_DATA / "case7_tplgy.m")
    components = parsed.components_by_block

    buses = {bus.number: bus for bus in components["DBAR"]}
    assert buses[6].bustype == ACBusTypes.ISOLATED
    assert len(buses) == 7

    lines = components["DLIN"]
    assert len(lines) == 8
    assert sum(line.available for line in lines) == 4
    assert all(not line.available for line in lines[:4])
    assert all(line.available for line in lines[4:])

    dc_links = components["DCLI"]
    assert len(dc_links) == 2


def test_parse_matpower_system_base_mva_override() -> None:
    parsed = parse_matpower_system(MATPOWER_DATA / "case5_dc.m", base_mva=200.0)
    constant = parsed.components_by_block["DCTE"][0]
    assert get_magnitude(constant.value) == pytest.approx(200.0)


def test_branch_classification() -> None:
    line_row = {
        "F_BUS": 1,
        "T_BUS": 2,
        "BR_R": 0.01,
        "BR_X": 0.1,
        "BR_B": 0.05,
        "RATE_A": 100.0,
        "TAP": 0.0,
        "SHIFT": 0.0,
        "BR_STATUS": 1,
    }
    transformer_row = {**line_row, "BR_R": 0.0, "TAP": 1.05}
    phase_shifter_row = {**line_row, "BR_R": 0.0, "SHIFT": 5.0}
    switch_row = {**line_row, "BR_R": 0.0, "BR_X": 0.0, "BR_B": 0.0, "RATE_A": 0.0}

    line = Branch([line_row])[0]
    transformer = Branch([transformer_row])[0]
    phase_shifter = Branch([phase_shifter_row])[0]
    switch = Branch([switch_row])[0]

    assert isinstance(line, ACLine)
    assert isinstance(transformer, TransformerDevice)
    assert isinstance(phase_shifter, PhaseShiftingTransformer)
    assert isinstance(switch, SwitchDevice)
    assert branch_block(line) == "DLIN"
    assert branch_block(transformer) == "DLIN_TRANSFORMER"
    assert branch_block(phase_shifter) == "DLIN_PHASE_SHIFT"
    assert branch_block(switch) == "DLIN_SWITCH"


def test_newton_raphson_solves_case5_dc() -> None:
    parsed = parse_matpower_system(MATPOWER_DATA / "case5_dc.m")
    solved = NewtonRaphsonPowerFlow(system=parsed.system, max_iterations=30, max_control_passes=0)
    results = solved.power_flow_results
    assert results.information.converged is True
    assert results.information.total_load_mw == pytest.approx(1015.0)
    assert results.information.solved_generation_mw == pytest.approx(1020.113, rel=1e-3)
    assert results.information.line_flow_overloads == 0


def test_newton_raphson_solves_large_activitys_case() -> None:
    parsed = parse_matpower_system(MATPOWER_DATA / "case_ACTIVSg10k.m")
    solved = NewtonRaphsonPowerFlow(system=parsed.system, max_iterations=30, max_control_passes=0)
    information = solved.power_flow_results.information

    assert information.converged is True
    assert information.solver == "newton-raphson"
    assert information.bus_count == 10000
    assert information.max_mismatch_pu <= 0.001


def test_fast_decoupled_solves_large_activitys_case_via_fallback() -> None:
    parsed = parse_matpower_system(MATPOWER_DATA / "case_ACTIVSg10k.m")
    solved = FastDecoupledPowerFlow(system=parsed.system, max_iterations=30, max_control_passes=0)
    information = solved.power_flow_results.information

    assert information.converged is True
    assert information.solver == "fast-decoupled"
    assert information.bus_count == 10000
    assert information.max_mismatch_pu <= 0.001


def test_newton_raphson_solves_large_activitys70k_case_with_shunts() -> None:
    parsed = parse_matpower_system(MATPOWER_DATA / "case_ACTIVSg70k.m")
    solved = NewtonRaphsonPowerFlow(system=parsed.system, max_iterations=30, max_control_passes=0)
    information = solved.power_flow_results.information

    assert information.converged is True
    assert information.solver == "newton-raphson"
    assert information.bus_count == 70000
    assert information.max_mismatch_pu <= 0.001
    assert information.solved_generation_mw == pytest.approx(612959.4, rel=1e-3)
    assert information.branch_active_losses_mw == pytest.approx(18243.0, rel=0.05)
    assert information.power_balance_mw == pytest.approx(0.0, abs=1e-3)


def test_acbuses_parses_matpower_bus_shunt_bs() -> None:
    from o_grid.models.m_models.topology import ACBuses

    buses, _ = ACBuses(
        [
            {
                "BUS_I": 26989.0,
                "BUS_TYPE": 1.0,
                "PD": 0.0,
                "QD": 0.0,
                "GS": 0.0,
                "BS": -1599.68,
                "BUS_AREA": 22.0,
                "VM": 1.039,
                "VA": -50.26,
                "BASE_KV": 765.0,
                "VMAX": 1.1,
                "VMIN": 0.9,
            },
            {
                "BUS_I": 2.0,
                "BUS_TYPE": 1.0,
                "PD": 0.0,
                "QD": 0.0,
                "GS": 0.0,
                "BS": 0.0,
                "BUS_AREA": 22.0,
                "VM": 1.0,
                "VA": 0.0,
                "BASE_KV": 765.0,
                "VMAX": 1.1,
                "VMIN": 0.9,
            },
        ],
        base_mva=100.0,
    )

    assert buses[0].capacitor_reactor.magnitude == pytest.approx(-1599.68)
    assert buses[1].capacitor_reactor.magnitude == pytest.approx(0.0)


def test_optimization_solves_case5_dc() -> None:
    parsed = parse_matpower_system(MATPOWER_DATA / "case5_dc.m")
    solved = OptimizationACPowerFlow(
        system=parsed.system,
        objective_function="squared_generation",
        max_iterations=100,
        print_iterations=False,
    )
    results = solved.power_flow_results
    assert results.information.converged is True
    assert results.information.solved_generation_mw == pytest.approx(1020.113, rel=1e-3)


def test_export_solution_excel(tmp_path: Path) -> None:
    from o_grid.exporter import SHEET_HEADERS

    parsed = parse_matpower_system(MATPOWER_DATA / "case5_dc.m")
    solved = NewtonRaphsonPowerFlow(parsed.system, max_control_passes=0)
    output_path = tmp_path / "case5_dc_solution.xlsx"

    export = ExportSolution(system=solved, format="excel", output_path=output_path)

    workbook = load_workbook(output_path, data_only=True)
    assert export.system is solved
    assert workbook.sheetnames == list(SHEET_HEADERS)
    summary = dict(workbook["Summary"].values)
    assert summary["Converged"] is True
    assert summary["Buses"] == 5


def test_plugin_parses_matpower_file(data_folder: Path) -> None:
    plugin = MatPowerParser()
    plugin._ctx = PluginContext(
        config=MatpowerConfig(
            pwf_path=str(data_folder / "mat" / "case5_dc.m"), system_name="plugin-matpower"
        )
    )

    result = plugin.on_build()

    assert result.is_ok()
    system = result.unwrap()
    assert isinstance(system, System)
    assert system.name == "plugin-matpower"
    assert plugin.ctx.metadata["parsed"].system is system


def test_plugin_without_pwf_path_returns_err() -> None:
    plugin = MatPowerParser()
    plugin._ctx = PluginContext(config=MatpowerConfig())

    result = plugin.on_build()

    assert result.is_err()
    assert "pwf_path" in result.unwrap_err()


def test_require_config_rejects_foreign_config() -> None:
    plugin = MatPowerParser()
    plugin._ctx = PluginContext(config=AnaredeConfig())

    with pytest.raises(ValueError, match="requires a MatpowerConfig"):
        plugin._require_config()


def test_parser_class_exposes_component_classes() -> None:
    parser = MatpowerInfrasysParser()
    parsed = parser.parse(MATPOWER_DATA / "case5_dc.m")
    assert parsed.component_classes["DBAR"] is ACBus
    assert parsed.component_classes["DLIN"] is ACLine
